from openpyxl import load_workbook
from openpyxl import Workbook
import requests
import json
import sys,os
# # import getpass

# def write_to_excel():
#     print(os.path)
#     wb = load_workbook('F:\\autoTest_workspace\\python_code\\e2e-test\\test_data\\写入数据.xlsx')
#     sheet = wb.get_sheet_by_name('已创建的公司')
#     sheet['A4'] = '来连连看'
#     wb.save(os.path.dirname(__file__) + '/../test_data/' +'写入数据.xlsx')
#     print('write success')

# def read_excel():
#     wb = load_workbook('写入数据.xlsx')
#     sheet = wb.get_sheet_by_name('已创建的公司')
#     print(sheet['A2'].value)

# # write_to_excel()
# read_excel()
def get_voucher_api(pageIndex):
    checkJournalList = []
    for i in range(1,17):
        url = 'https://api-firms.guanplus.com/api/v1/journal/search?year=2017&month=8&keyword=&pageIndex=' + str(i) + '&pageSize=10&type='
        headers = {
                'Accept':'application/json',
                'Accept-Encoding':'gzip, deflate, sdch',
                'Accept-Language':'zh-CN,zh;q=0.8',
                'Cache-Control':'no-cache',
                'Connection':'keep-alive',
                'Content-Type':'application/json',
                'Authorization':'bearer aOmvOukO67OHdvqLWbnTyULE_eOnsPqZqnDxes2kgYQ1owKRGVoINwe85CCf5Ns2iw5bKBUjmL6kb-F0b6W8uMo58d4myuDINR35fBKQq-w4wb-l4JgIg2aqEMH1qid34VA_-wLVzZztjnrVsIHXYKVLmT58l33mKrU3E2o_7VQWP-qRuc_zVdH_LTWRv93uwzfOz3PhrkueXRmBT_uZkIb1Kl-ESyCnLKzfbvzmNsJ1CdFTlfvdRAf3cvQI_ZKDNNHLGRfsiZzzWB7GxCcy8B1g1aM',
                'company_id':'5a8a5e65-0f17-4353-815a-9046199c6f5b',
                'accountbook_id':'94abec6d-1afe-4c2f-bc26-88043fe4754d',
                'Host':'api-firms.guanplus.com',
                'Origin':'https://firms.guanplus.com',
                'Referer':'//firms.guanplus.com/app/finance/voucher',
                'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.221 Safari/537.36 SE 2.X MetaSr 1.0'
            }
        request = requests.get(url,headers=headers)
        resposeData = json.loads(request.text)
        journalList = resposeData['list']
        
        for i in journalList:
            journalEntryLineItemList = i['journalEntryLineItemModels']
            for j in journalEntryLineItemList:
                checkJournalItemDict = {}
                checkJournalItemDict['journalNumber'] = i['journalEntryNumber']
                checkJournalItemDict['accountCode'] = j['account']['id']
                checkJournalItemDict['accountName'] = j['account']['name']
                checkJournalItemDict['dcDirection'] = j['debitCreditType']
                checkJournalItemDict['amount'] = str(j['amount'])
                checkJournalList.append(checkJournalItemDict)
    print(checkJournalList)
            
    wb = load_workbook('F:\\autoTest_workspace\\python_code\\e2e-test\\test_data\\凭证.xlsx')
    sheet = wb.get_sheet_by_name('说明')
    for lists,i in zip(checkJournalList,range(2,(len(checkJournalList) + 2))):
        rowNum = i
        sheet['A' + str(rowNum)] = lists['journalNumber'] 
        sheet['B' + str(rowNum)] = lists['accountCode']
        sheet['C' + str(rowNum)] = lists['accountName']
        sheet['D' + str(rowNum)] = lists['dcDirection']
        sheet['E' + str(rowNum)] = lists['amount']


    wb.save('凭证.xlsx')
    print('write success')
    # return checkJournalList

get_voucher_api('1')
def write_to_excel():
    wb = load_workbook('F:\\autoTest_workspace\\python_code\\e2e-test\\test_data\\凭证.xlsx')
    sheet = wb.get_sheet_by_name('（一般纳税人）流水单生成凭证校验')
    sheet['A4'] = '来连连看'
    wb.save('凭证.xlsx')
    print('write success')
    


